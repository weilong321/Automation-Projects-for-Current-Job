from selenium.webdriver.chrome.webdriver import WebDriver as Chrome
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from time import sleep

from openpyxl import load_workbook
from openpyxl import utils
import pandas as pd

import os
import shutil


# Get ascii files
def get_df_ledgers(path):
    df = pd.read_excel(path, sheet_name="3161040000 Trade Debtors") # change this
    GL_account = df.iat[0, 0]
    df = df.drop(df.index[:2])
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])
    df = df.iloc[:, :8]
    df.columns = ["Supplier No", "Supplier Name", "Currency", "Oct and prior", "Nov-22", "Dec-22", "Jan-23", "Total"] # change this
    df = df.iloc[:-5].reset_index(drop=True).fillna(0)
    df = df[(df["Oct and prior"] != 0) & (df["Supplier Name"] != "IATA BSP")].sort_values(by="Supplier No").reset_index(drop=True)
    folder_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(GL_account)
    SL_account_list = df[df['Oct and prior'] != 0]['Supplier No'].tolist()
    if not os.path.exists(folder_path):
        return GL_account, SL_account_list, df
    file_names = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    temp = [x for x in file_names if x not in SL_account_list]
    temp2 = [int(x.split(".")[0].split("_")[1]) for x in temp]
    remaining_sl = []
    for sl in SL_account_list:
        if sl not in temp2:
            remaining_sl.append(sl)
    return GL_account, remaining_sl, df
def find_type_click_element(xpath, not_button=None, input=None, needs_enter=None, pause=None, wait=True, style=None):
    if wait:
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
    element = driver.find_element(By.XPATH, xpath)
    if pause:
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        action = ActionChains(driver)
        action.move_to_element(element).click().perform()
        sleep(1)
    else:
        element.click()
    if not_button:
        element.clear()
        element.send_keys(input)
    if needs_enter:
        element.send_keys(Keys.ENTER)
    return
def rename_move_file(general, subledger, is_first=True, to_be_updated=False, multiple=False, is_temp=False):
    desktop_path = ""
    initial_path = ""
    if is_first:
        is_first = False
        sleep(2)
    if to_be_updated:
        desktop_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(general)
    else:
        # desktop_path = "C:/Users/ctm_mchen/OneDrive - Helloworld Travel Ltd/Desktop/General_Ledger_{}/".format(general)
        desktop_path = "C:/Users/Matthew Chen/Desktop/subledger/General_Ledger_{}/".format(general)
    if is_temp:
        desktop_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_CSV_Files/".format(general)
        initial_path = "C:/temp/"
    else:
        # initial_path = "C:/Users/ctm_mchen/Downloads"
        initial_path = "C:/Users/Matthew Chen/Downloads"
    if not os.path.exists(desktop_path):
        os.mkdir(desktop_path)
    if multiple:
        agm_files = [f for f in os.listdir(initial_path) if f.endswith(".agm")]
        agm_filename = os.path.splitext(agm_files[0])[0]
        pdf_file = agm_filename + ".pdf"
        if pdf_file in os.listdir(initial_path):
            os.remove(os.path.join(initial_path, pdf_file))
        shutil.move(os.path.join(initial_path, agm_files[0]),
                    os.path.join(desktop_path, "Subledger_{}.agm".format(subledger))) 
    else:
        filename = max([initial_path + "/" + f for f in os.listdir(initial_path)],key=os.path.getctime)
        ext = os.path.splitext(filename)[1]
        shutil.move(filename,os.path.join(desktop_path,"Subledger_{}{}".format(subledger, ext)))
def initiate_driver():
    global driver
    options = Options()
    prefs = {'profile.default_content_setting_values.automatic_downloads': 1}
    options.add_experimental_option("prefs", prefs)
    driver = Chrome("chromedriver.exe", options=options)
    driver.maximize_window()
    driver.get(WEBSITE)
    WEBSITE_login()
    if apx_or_qbt == "QBT":
        get_token()
def get_token():
    find_type_click_element(REQUEST_NEW_TOKEN_XPATH)
    logon_key = input("Type in logon token here: \n")
    find_type_click_element(LOGON_TOKEN_XPATH, not_button=True, input=logon_key)
    find_type_click_element(OK_XPATH)
    return
def WEBSITE_login():
    find_type_click_element(USER_XPATH, not_button=True, input=USERNAME)
    find_type_click_element(PASS_XPATH, not_button=True, input=PASSWORD)
    find_type_click_element(LOGON_XPATH)
    return
def download_agm_file(general, subledger):
    try:
        if apx_or_qbt == "QBT":
            find_type_click_element(MENU_XPATH)
            find_type_click_element(BACK_OFFICE_XPATH)
            find_type_click_element(OPEN_ITEMS_XPATH)
            find_type_click_element(OPEN_ITEMS_REPORT)
            find_type_click_element(OPEN_ITEMS_BY_REFERENCE)
            find_type_click_element(OPEN_ITEMS_GENERAL)
            find_type_click_element(ACCOUNTANCY_PERIOD, not_button=True, input="202307")
            find_type_click_element(GL_ACCOUNT1, not_button=True, input=general)
            find_type_click_element(GL_ACCOUNT2, not_button=True, input=general)
            find_type_click_element(SL_ACCOUNT1, not_button=True, input=subledger)
            find_type_click_element(SL_ACCOUNT2, not_button=True, input=subledger)
            find_type_click_element(SL_TYPE1, not_button=True, input="1")
            find_type_click_element(SL_TYPE2, not_button=True, input="1")
            find_type_click_element(VALIDATE)
            find_type_click_element(OK)
            find_type_click_element(EMAIL)
            find_type_click_element(CREATE_ASCII)
            find_type_click_element(VALIDATE2, pause=True)
            find_type_click_element(OK2, pause=True)
            wait = WebDriverWait(driver, 120)
            wait.until(EC.invisibility_of_element_located((By.XPATH, OK2)))
            rename_move_file(general, subledger, to_be_updated=True, multiple=True)
        elif apx_or_qbt == "APX":
            find_type_click_element(MENU_XPATH)
            find_type_click_element(BACK_OFFICE_XPATH)
            find_type_click_element(OPEN_ITEMS_XPATH)
            find_type_click_element(OPEN_ITEMS_REPORT)
            find_type_click_element(OPEN_ITEMS_BY_SUBLEDGER_ONLY)
            find_type_click_element(OPEN_ITEMS_GENERAL)
            find_type_click_element(ACCOUNTANCY_PERIOD, not_button=True, input="202307")
            find_type_click_element(SL_TYPE, not_button=True, input="2")
            find_type_click_element(SL_ACCOUNT1, not_button=True, input=subledger)
            find_type_click_element(SL_ACCOUNT2, not_button=True, input=subledger)
            find_type_click_element(LAUNCHED_ENTRIES)
            find_type_click_element(SELECT_GL_ACCOUNT)
            find_type_click_element(RETROSPECTIVE)
            find_type_click_element(VALIDATE)
            find_type_click_element(OK)
            find_type_click_element(EMAIL)
            find_type_click_element(CREATE_ASCII)
            find_type_click_element(VALIDATE2, pause=True)
            find_type_click_element(OK2, pause=True)
            wait = WebDriverWait(driver, 120)
            wait.until(EC.invisibility_of_element_located((By.XPATH, OK2)))
            rename_move_file(general, subledger, to_be_updated=True, multiple=True)
        return True
    except:
        return False
def get_updated_ascii(general, subledgers):
    initiate_driver()
    for subledger in subledgers:
        while True:
            if download_agm_file(general, subledger):
                break
    return

# read ascii files and get combined df
def append_all(dfs, names, df, name):
    dfs.append(df)
    names.append(name)
    return
def read_agm_csv(filepath):
    ls = [0, 0, 0, 0, 0, 0, 0, 0, 0]
    with open(filepath, 'r') as file:
        for line in file:
            if line.startswith('D3'):
                temp = line.strip().split("\x1f")
                fy = temp[3]
                amount = float(temp[12].strip().replace(",", ""))
                if fy.startswith("2017"):
                    ls[0] += amount
                    ls[8] += amount
                elif fy.startswith("2018"):
                    ls[1] += amount
                    ls[8] += amount
                elif fy.startswith("2019"):
                    ls[2] += amount
                    ls[8] += amount
                elif fy.startswith("2020"):
                    ls[3] += amount
                    ls[8] += amount
                elif fy.startswith("2021"):
                    ls[4] += amount
                    ls[8] += amount
                elif fy.startswith("20220"):
                    ls[5] += amount
                    ls[8] += amount
                elif fy.startswith("20221"):
                    ls[6] += amount
                    ls[8] += amount
                elif fy == "202301" or fy == "202302" or fy == "202303" or fy == "202304":
                    ls[7] += amount
                    ls[8] += amount
    return ls
def get_combined_df(main_df, general):
    folder_path = "C:/Users/Matthew Chen/Desktop/subledger/{}_AGM_Files/".format(general)
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    files = sorted(files, key=lambda x: len(os.path.basename(x)))
    name_ls = []
    ls = []
    for file in files:
        name = file.split(".")[0].split("_")[1]
        path = folder_path + file
        name_ls.append(name)
        ls.append(read_agm_csv(path))
    fy_df = pd.DataFrame(ls, columns=['FY17', 'FY18', 'FY19', 'FY20', 'FY21', 'FY22 (to Mar22)', 'FY22 (Apr-Jun22)', 'FY23 (Jul-Oct22)', 'Total (Oct and prior)'])
    temp = main_df.sort_values(by=['Supplier No']).reset_index(drop=True)
    combined = pd.concat([temp[['Supplier No', 'Supplier Name', 'Oct and prior']], fy_df], axis=1).fillna(0)
    combined = combined.assign(Diff = combined['Oct and prior'] - combined['Total (Oct and prior)'])
    append_all(dfs, names, combined, "Combined Summary")
    return

# add to workbook
def add_sheets(path):
    wb = load_workbook(path)
    ws = None
    for i in range(len(names)):
        ws = wb.create_sheet(names[i])
        for col_idx, value in enumerate(dfs[i].columns.tolist()):
            ws.cell(row=1, column=col_idx+1).value = value
        for row_idx, row in dfs[i].iterrows():
            for col_idx, cell_value in enumerate(row.tolist()):
                try:
                    rounded_value = round(float(cell_value), 2)
                    ws.cell(row=row_idx+2, column=col_idx+1).value = rounded_value
                except:
                    ws.cell(row=row_idx+2, column=col_idx+1).value = cell_value
        ws.freeze_panes = 'A2'
        for column in ws.columns:
            for cell in column:
                column_letter = utils.get_column_letter(cell.column)
                ws.column_dimensions[column_letter].auto_size = True
    wb.save(path)
    dfs.clear()
    names.clear()
    return

global apx_or_qbt
while True:
    apx_or_qbt = input("APX or QBT: ")
    if apx_or_qbt == "QBT":
        from qbt import *
        break
    elif apx_or_qbt == "APX":
        from apxnz import *
        break
    else:
        print("Invalid input, try again.")
        continue

global dfs
global names
dfs = []
names = []
path = "APX Debtors Jan23.xlsx"

general, subledgers, main_df = get_df_ledgers(path)
get_updated_ascii(general, subledgers)
combined = get_combined_df(main_df, general)
add_sheets(path)